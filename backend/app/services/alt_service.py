import asyncio
import json
from pathlib import Path
from fastapi import Request
from typing import Dict, Any, List # Added List
import openpyxl # Added openpyxl

from app.core.storage import storage_service
from app.models.schemas import SSEProgress, SSEPartial, SSEDone, SSEMetadata # Added SSEMetadata

class AltService:
    def __init__(self):
        self.jobs: Dict[str, Dict[str, Any]] = {}

    def register_job(self, job_id: str, file_path: Path):
        self.jobs[job_id] = {"file_path": file_path}

    def get_job_filepath(self, job_id: str) -> Path | None:
        job = self.jobs.get(job_id)
        return job.get("file_path") if job else None

    async def process_file(self, request: Request, file_path: Path):
        try:
            yield {
                "event": "progress",
                "data": SSEProgress(percent=10, message="檔案讀取完成，開始分析...").model_dump_json()
            }
            await asyncio.sleep(0.7)

            # --- 新增：讀取 Excel 檔案並提取查詢欄位和目標 ---
            query_fields: List[str] = []
            query_targets: List[str] = []
            try:
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active

                # 提取第一列作為查詢欄位 (A1, A2, A3...) 
                if sheet.max_column >= 1:
                    for row_idx in range(1, sheet.max_row + 1):
                        cell_value = sheet.cell(row=row_idx, column=1).value
                        if cell_value is not None:
                            query_fields.append(cell_value)

                # 提取第一行（從第二個儲存格開始）作為查詢目標 (B1, C1, D1...)
                if sheet.max_row >= 1:
                    query_targets = [cell.value for cell in sheet[1][1:] if cell.value is not None]

                # 發送 metadata 事件給前端
                yield {
                    "event": "metadata",
                    "data": SSEMetadata(query_fields=query_fields, query_targets=query_targets).model_dump_json()
                }
                await asyncio.sleep(0.5)

            except Exception as e:
                print(f"Error reading Excel file: {e}")
                yield {
                    "event": "error",
                    "data": json.dumps({"message": f"讀取 Excel 檔案失敗: {e}"})
                }
                return # 讀取失敗則終止處理
            # --- 結束新增 ---

            mock_text = "正在比對第一行資料... 找到可能替代品 A... 相似度 92%... 正在驗證庫存..."
            for char in mock_text:
                if await request.is_disconnected():
                    print(f"Client for job {request.url.path} disconnected, stopping SSE stream.")
                    break
                yield {
                    "event": "partial",
                    "data": SSEPartial(text=char).model_dump_json()
                }
                await asyncio.sleep(0.05)
            
            yield {
                "event": "progress",
                "data": SSEProgress(percent=50, message="初步比對完成，進行深度掃描...").model_dump_json()
            }
            await asyncio.sleep(1)

            mock_text_2 = "\n深度掃描發現替代品 B，成本降低 15%... 正在產生報告..."
            for char in mock_text_2:
                if await request.is_disconnected():
                    print(f"Client for job {request.url.path} disconnected, stopping SSE stream.")
                    break
                yield {
                    "event": "partial",
                    "data": SSEPartial(text=char).model_dump_json()
                }
                await asyncio.sleep(0.05)

            yield {
                "event": "progress",
                "data": SSEProgress(percent=90, message="報告產生中...").model_dump_json()
            }
            await asyncio.sleep(0.5)

            # 使用原始上傳檔案作為下載連結
            download_url = storage_service.make_downloadable(file_path)
            
            yield {
                "event": "done",
                "data": SSEDone(download_url=download_url).model_dump_json()
            }

        except asyncio.CancelledError:
            print(f"SSE stream for job {request.url.path} was cancelled.")
        except Exception as e:
            print(f"Error during SSE processing for job {request.url.path}: {e}")
            yield {
                "event": "error",
                "data": json.dumps({"message": "處理過程中發生錯誤"})
            }
        finally:
            pass